import os, json, datetime
import pandas as pd
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.contrib.auth.models import Group
from .forms import Register
from .models import User
from audit.models import AuditPlan, CheckList, Report  # Asegúrate que "audit" sea tu app correcta
from django.views.decorators.csrf import csrf_exempt
from django.utils.translation import gettext as _
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.merge import CellRange
from openpyxl import Workbook


# Vista de registro
def register_view(request):
    if request.method == 'POST':
        form = Register(request.POST)
        if form.is_valid():
            user = form.save()
            # Si el usuario es superUser, establecer los permisos de superusuario
            if user.role == 'superUser':
                user.is_staff = True
                user.is_superuser = True
                user.save()
            login(request, user)
            messages.success(request, _(f'Registro correcto. Tu código de verificación es: {user.token}'))
            return redirect('verify')
        else:
            username = form.cleaned_data.get('username')
            if User.objects.filter(username=username).exists():
                messages.warning(request, _('El nombre de usuario ya está en uso.'))

            email = form.cleaned_data.get('email')
            if User.objects.filter(email=email).exists():
                messages.warning(request, _('El correo electrónico ya está registrado.'))

            password1 = form.cleaned_data.get('password1')
            password2 = form.cleaned_data.get('password2')
            if password1 != password2:
                messages.warning(request, _('Las contraseñas no coinciden.'))
    else:
        form = Register()
    return render(request, 'register.html', {'form': form})

def verify_user(request):
    if request.method == 'POST':
        token = request.POST.get('token')
        try:
            user = User.objects.get(token=token)
            user.is_active = True
            user.save()
            messages.success(request, _('Usuario verificado exitosamente. Ya puedes iniciar sesión.'))
            return redirect('login')
        except User.DoesNotExist:
            messages.error(request, _('Código inválido.'))
    return render(request, 'verify.html')

# Vista de login
def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            nombre = form.cleaned_data.get('username')
            contraseña = form.cleaned_data.get('password')
            codigo = form.cleaned_data.get('code')
            user = authenticate(request, username=nombre, password=contraseña, code=codigo)
            if user:
                messages.success(request, _("Inicio de sesión correcto"))
                login(request, user)
                return redirect('profile')
        else:
            messages.warning(request, _('Los datos son incorrectos'))
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

# Vista de perfil
def profile_view(request):
    if request.user.is_authenticated:
        return render(request, "profile.html")
    else:
        messages.info(request, _("Inicia sesión primero"))
        return redirect('login')

# Vista de logout
def logout_view(request):
    if request.method == 'POST':
        logout(request)
        messages.success(request, _("Cierre de sesión correcto"))
        return redirect('register')
    return redirect('register')

# Vista del Dashboard con edición y eliminación inline
def admin_dashboard(request):
    if request.user.role not in ['auditLeaderUser', 'superUser', 'organizationUser']:
        return JsonResponse({"error": "No autorizado"}, status=403)
    # if not request.user.is_authenticated:
    #     messages.warning(request, _("Debes iniciar sesión para ver el dashboard."))
    #     return redirect('login')

    org =  request.user.organization

    # Procesamiento de formularios POST
    if request.method == 'POST':
        form_type = request.POST.get('form_type')

        if form_type == 'edit_user':
            user_id = request.POST.get('user_id')
            user = get_object_or_404(User, id=user_id, organization=org)
            user.first_name = request.POST.get('first_name', user.first_name)
            user.last_name = request.POST.get('last_name', user.last_name)
            user.email = request.POST.get('email', user.email)
            user.role = request.POST.get('role', user.role)
            user.save()
            messages.success(request, _("Usuario %(username)s actualizado.") % {"username": user.username})


        elif form_type == 'delete_user':
            user_id = request.POST.get('user_id')
            user = get_object_or_404(User, id=user_id, organization=org)
            user.delete()
            messages.success(request, _(f"Usuario eliminado."))

        elif form_type == 'create_user':
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            role = request.POST.get('role')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')

            if User.objects.filter(username=username).exists():
                messages.error(request, f"El usuario {username} ya existe.")
            else:
                new_user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    role=role,
                    organization_id=org
                )
                # Si el usuario es superUser, establecer los permisos de superusuario
                if role == 'superUser':
                    new_user.is_staff = True
                    new_user.is_superuser = True
                    new_user.save()
                messages.success(request, f"Usuario {username} creado correctamente.")

        return redirect('admin_dashboard')

    context = {
        'users': User.objects.filter(organization=org),
        'audit_plans': AuditPlan.objects.filter(organization=org),
        'checklists': CheckList.objects.filter(organization=org).select_related("leader_auditor", "organization"),
        'reports': Report.objects.filter(organization=org),
        'groups': Group.objects.all(),
    }
    return render(request, 'dashboard.html', context)

# @csrf_exempt
def download_excel_checklist(request):
    if request.user.role not in ['auditLeaderUser', 'superUser', 'organizationUser']:
        return JsonResponse({"error": "No autorizado"}, status=403)

    checklist_id = request.GET.get("checklist_id")
    if not checklist_id:
        return JsonResponse({"error": "ID de checklist no proporcionado"}, status=400)

    try:
        checklist = CheckList.objects.get(id=checklist_id)
        audit_data = checklist.audit_data

        if isinstance(audit_data, str):
            audit_data = json.loads(audit_data)

        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Hoja 1: Auditoría
            df = pd.DataFrame(audit_data)
            df.to_excel(writer, index=False, sheet_name="Auditoría", startrow=1)
            ws = writer.sheets["Auditoría"]

            header_fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

            for col_num, column_cells in enumerate(ws.iter_cols(min_row=2, max_row=2), 1):
                cell = column_cells[0]
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                ws.column_dimensions[cell.column_letter].width = 30

            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    cell.border = border
                ws.row_dimensions[row[0].row].height = 80

            # Hoja 2: Resumen
            meta_data = {
                "Proceso": [checklist.process],
                "Lugar": [checklist.place],
                "Organización": [checklist.organization.name],
                "Auditor Líder": [checklist.leader_auditor.first_name],
                "Tipo de Proceso": [checklist.process_type],
                "Cláusulas": [checklist.clauses_list]
            }
            meta_df = pd.DataFrame(meta_data)
            meta_df.to_excel(writer, index=False, sheet_name="Resumen")

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=checklist_{checklist.id}.xlsx'
        return response

    except CheckList.DoesNotExist:
        return JsonResponse({"error": "Checklist no encontrada"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    
# @csrf_exempt
def download_excel_audit_plan(request):
    # autorización
    if request.user.role not in ['auditLeaderUser', 'superUser', 'organizationUser']:
        return JsonResponse({"error": "No autorizado"}, status=403)

    plan_id = request.GET.get("plan_id")
    if not plan_id:
        return JsonResponse({"error": "ID de plan no proporcionado"}, status=400)

    try:
        plan = AuditPlan.objects.get(id=plan_id)
        content = plan.plan_content or {}
        if isinstance(content, str):
            content = json.loads(content)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            wb = writer.book
            # eliminar hoja por defecto
            if wb.sheetnames:
                wb.remove(wb[wb.sheetnames[0]])

            # estilos comunes
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            wrap = Alignment(wrap_text=True, vertical="top")
            header_fill = PatternFill("solid", fgColor="999999")
            section_fill = PatternFill("solid", fgColor="C0C0C0")
            white_bold = Font(color="FFFFFF", bold=True)
            black_bold = Font(color="000000", bold=True)

            def set_cols(ws, widths):
                for i, w in enumerate(widths, start=1):
                    ws.column_dimensions[get_column_letter(i)].width = w

            # === 1) Portada ===
            ws0 = wb.create_sheet("Portada")
            set_cols(ws0, [3, 40, 40])
            # Título principal combinado B2:D4
            ws0.merge_cells("B2:D4")
            c = ws0["B2"]
            c.value = plan.organization.name.upper()
            c.font = Font(size=20, bold=True)
            c.alignment = center

            # === 2) Alcance – Criterios – Objetivos ===
            ws1 = wb.create_sheet("Alcance - Criterios - Objetivos")
            set_cols(ws1, [3, 34, 58, 26])

            # 2.1 Participantes
            ws1.merge_cells("B2:B4")
            c = ws1["B2"]
            c.border = border
            c.value = plan.organization.name.upper()
            c.font = Font(size=20, bold=True)
            c.alignment = center
            c = ws1["B3"]
            c.border = border
            c = ws1["B4"]
            c.border = border

            ws1.merge_cells("C2:C4")
            c = ws1["C2"]
            c.value = "Name".upper()
            c.font = Font(size=20, bold=True)
            c.alignment = center
            c.border = border
            c = ws1["C3"]
            c.border = border
            c = ws1["C4"]
            c.border = border

            ws1.merge_cells("D2:D4")
            c = ws1["D2"]
            c.value = "Fechas".upper()
            c.font = Font(size=20, bold=True)
            c.alignment = center
            c.border = border
            c = ws1["D3"]
            c.border = border
            c = ws1["D4"]
            c.border = border

            ws1.merge_cells("B6:D6")
            c = ws1["B6"]
            c.value = "Participantes en la elaboración del Programa de Auditoría"
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            headers = ["Rol", "Nombre", "Asistencia"]
            for j, txt in enumerate(headers, start=2):
                cell = ws1.cell(row=7, column=j, value=txt)
                cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            for i, row_data in enumerate(content.get("tabla-revision", []), start=8):
                ws1.row_dimensions[i].height = 50
                for j, val in enumerate(row_data, start=2):
                    cell = ws1.cell(row=i, column=j, value=val)
                    cell.alignment = wrap; cell.border = border

            # espacio
            sep = ws1.max_row + 2

            # 2.2 Alcance
            ws1.merge_cells(start_row=sep, start_column=2, end_row=sep, end_column=4)
            c = ws1.cell(row=sep, column=2, value="Alcance del Programa de Auditoría")
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            headers2 = ["Tipo", "Descripción", "Observaciones"]
            for j, txt in enumerate(headers2, start=2):
                cell = ws1.cell(row=sep+1, column=j, value=txt)
                cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            for k, row_data in enumerate(content.get("tabla-alcance", []), start=sep+2):
                ws1.row_dimensions[k].height = 40
                for j, val in enumerate(row_data, start=2):
                    cell = ws1.cell(row=k, column=j, value=val)
                    cell.alignment = wrap; cell.border = border

            # 2.3 Criterios
            sep2 = ws1.max_row + 2
            ws1.merge_cells(start_row=sep2, start_column=2, end_row=sep2, end_column=4)
            c = ws1.cell(row=sep2, column=2, value="Criterios del Programa de Auditoría")
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            headers3 = ["Tipo", "Descripción", "Observaciones"]
            for j, txt in enumerate(headers3, start=2):
                cell = ws1.cell(row=sep2+1, column=j, value=txt)
                cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            for k, row_data in enumerate(content.get("tabla-criterios", []), start=sep2+2):
                ws1.row_dimensions[k].height = 40
                for j, val in enumerate(row_data, start=2):
                    cell = ws1.cell(row=k, column=j, value=val)
                    cell.alignment = wrap; cell.border = border

            # === 3) Incertidumbre ===
            ws2 = wb.create_sheet("Uncertainty")
            set_cols(ws2, [3, 20, 50, 50, 27])

            # 3.1 Oportunidades
            ws2.merge_cells("B6:E6")
            c = ws2["B6"]
            c.value = "Determinación y Evaluación de Oportunidades"
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            headers4 = ["Oportunidad", "Descripción", "Acciones", "Recursos"]
            for j, txt in enumerate(headers4, start=2):
                cell = ws2.cell(row=7, column=j, value=txt)
                cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            for i, row_data in enumerate(content.get("tabla-oportunidades", []), start=8):
                ws2.row_dimensions[i].height = 40
                for j, val in enumerate(row_data, start=2):
                    cell = ws2.cell(row=i, column=j, value=val)
                    cell.alignment = wrap; cell.border = border

            # espacio y Riesgos
            sep3 = ws2.max_row + 2
            ws2.merge_cells(start_row=sep3, start_column=2, end_row=sep3, end_column=5)
            c = ws2.cell(row=sep3, column=2, value="Determinación y Evaluación de Riesgos")
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            for j, txt in enumerate(headers4, start=2):
                cell = ws2.cell(row=sep3+1, column=j, value=txt)
                cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            for k, row_data in enumerate(content.get("tabla-riesgos", []), start=sep3+2):
                ws2.row_dimensions[k].height = 40
                for j, val in enumerate(row_data, start=2):
                    cell = ws2.cell(row=k, column=j, value=val)
                    cell.alignment = wrap; cell.border = border

            # === 4) Recursos ===
            ws3 = wb.create_sheet("Resources")
            set_cols(ws3, [3, 35, 50, 30])

            # Título
            ws3.merge_cells("B6:D6")
            c = ws3["B6"]
            c.value = "Recursos para la Gestión del Programa de Auditorías"
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            # Encabezados
            cell = ws3.cell(row=7, column=2, value="Tipo")
            cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            ws3.merge_cells(start_row=7, start_column=3, end_row=7, end_column=4)
            cell = ws3.cell(row=7, column=3, value="Descripción")
            cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            # Contenido
            for i, row_data in enumerate(content.get("tabla-recursos", []), start=8):
                ws3.row_dimensions[i].height = 40

                # Tipo en columna B
                cell = ws3.cell(row=i, column=2, value=row_data[0])
                cell.alignment = wrap; cell.border = border

                # Descripción en columnas C y D
                ws3.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
                cell = ws3.cell(row=i, column=3, value=row_data[1])
                cell.alignment = wrap; cell.border = border


            # espacio y Equipo Auditor
            sep4 = ws3.max_row + 2
            ws3.merge_cells(start_row=sep4, start_column=2, end_row=sep4, end_column=4)
            c = ws3.cell(row=sep4, column=2, value="Equipo Auditor")
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            headers6 = ["Nombre", "Rol", "Correo"]
            for j, txt in enumerate(headers6, start=2):
                cell = ws3.cell(row=sep4+1, column=j, value=txt)
                cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            for k, row_data in enumerate(content.get("tabla-equipoAuditor", []), start=sep4+2):
                ws3.row_dimensions[k].height = 40
                for j, val in enumerate(row_data, start=2):
                    cell = ws3.cell(row=k, column=j, value=val)
                    cell.alignment = wrap; cell.border = border

            # === 5) Implementación ===
            ws4 = wb.create_sheet("Implementation")
            set_cols(ws4, [3, 20, 20, 50, 20, 20])

            ws4.merge_cells("B6:F6")
            c = ws4["B6"]
            c.value = "Implementación del Programa de Auditoría"
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            headers7 = ["Etapa", "Fase", "Descripción", "Fecha Inicio", "Fecha Final"]
            for j, txt in enumerate(headers7, start=2):
                cell = ws4.cell(row=7, column=j, value=txt)
                cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            for i, row_data in enumerate(content.get("tabla-estapasAud", []), start=8):
                ws4.row_dimensions[i].height = 40
                for j, val in enumerate(row_data, start=2):
                    cell = ws4.cell(row=i, column=j, value=val)
                    cell.alignment = wrap; cell.border = border

            # espacio y Metodología
            sep5 = ws4.max_row + 2
            ws4.merge_cells(start_row=sep5, start_column=2, end_row=sep5, end_column=6)
            c = ws4.cell(row=sep5, column=2, value="Metodología de Auditoría")
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            # encabezados
            ws4.cell(row=sep5+1, column=2, value="Tipo").fill = section_fill
            ws4.cell(row=sep5+1, column=2).font = black_bold; ws4.cell(row=sep5+1, column=2).alignment = center; ws4.cell(row=sep5+1, column=2).border = border

            ws4.merge_cells(start_row=sep5+1, start_column=3, end_row=sep5+1, end_column=4)
            cell = ws4.cell(row=sep5+1, column=3, value="Descripción")
            cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            ws4.merge_cells(start_row=sep5+1, start_column=5, end_row=sep5+1, end_column=6)
            cell = ws4.cell(row=sep5+1, column=5, value="Observaciones")
            cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            # contenido
            for k, row_data in enumerate(content.get("tabla-metodologia", []), start=sep5+2):
                ws4.row_dimensions[k].height = 40

                # Tipo (columna B)
                cell = ws4.cell(row=k, column=2, value=row_data[0])
                cell.alignment = wrap; cell.border = border

                # Descripción (columnas C y D)
                ws4.merge_cells(start_row=k, start_column=3, end_row=k, end_column=4)
                cell = ws4.cell(row=k, column=3, value=row_data[1])
                cell.alignment = wrap; cell.border = border

                # Observaciones (columnas E y F)
                ws4.merge_cells(start_row=k, start_column=5, end_row=k, end_column=6)
                cell = ws4.cell(row=k, column=5, value=row_data[2])
                cell.alignment = wrap; cell.border = border


            # === 6) Plan de Auditoría ===
            ws5 = wb.create_sheet("Audit Plan")
            set_cols(ws5, [3,20,20,20,20,20,40,15,15,30,50])

            # Objetivos
            ws5.merge_cells("B6:L6")
            c = ws5["B6"]
            c.value = "Objetivos de Auditoría"
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            objs = content.get("tabla-objetivos", [])
            text = "\n".join(f"{i+1}. {o[1]}" for i,o in enumerate(objs))
            ws5.row_dimensions[7].height = 70
            cell = ws5.cell(row=7, column=2, value=text)
            cell.alignment = wrap; cell.border = border
            ws5.merge_cells("B7:L7")

            # Scope dentro de Plan
            sep6 = 8
            ws5.merge_cells(start_row=sep6, start_column=2, end_row=sep6, end_column=12)
            c = ws5.cell(row=sep6, column=2, value="Alcance")
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            # subtítulos y datos tabla-alcance
            # Encabezados
            for j, (txt, col_start, col_end) in enumerate([
                ("Tipo", 2, 4), ("Descripción", 5, 8), ("Observaciones", 9, 12)
            ], start=1):
                ws5.merge_cells(start_row=sep6+1, start_column=col_start, end_row=sep6+1, end_column=col_end)
                cell = ws5.cell(row=sep6+1, column=col_start, value=txt)
                cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            # Contenido
            for k, row_data in enumerate(content.get("tabla-alcance", []), start=sep6+2):
                ws5.row_dimensions[k].height = 40

                # Tipo → B-D
                ws5.merge_cells(start_row=k, start_column=2, end_row=k, end_column=4)
                cell = ws5.cell(row=k, column=2, value=row_data[0])
                cell.alignment = wrap; cell.border = border

                # Descripción → E-H
                ws5.merge_cells(start_row=k, start_column=5, end_row=k, end_column=8)
                cell = ws5.cell(row=k, column=5, value=row_data[1])
                cell.alignment = wrap; cell.border = border

                # Observaciones → I-L
                ws5.merge_cells(start_row=k, start_column=9, end_row=k, end_column=12)
                cell = ws5.cell(row=k, column=9, value=row_data[2])
                cell.alignment = wrap; cell.border = border


            # Observaciones (última columna) para cada fila de planAud
            sep7 = ws5.max_row + 2
            ws5.merge_cells(start_row=sep7, start_column=2, end_row=sep7, end_column=12)
            c = ws5.cell(row=sep7, column=2, value="Plan de Auditoría Detallado")
            c.fill = header_fill; c.font = white_bold; c.alignment = center; c.border = border

            headers9 = ["#", "Proceso", "Dependencia", "Lugar", "Método", "Cláusulas", "Fecha", "Hora", "Responsable", "Auditor", "Observaciones"]
            for j, txt in enumerate(headers9, start=2):
                cell = ws5.cell(row=sep7+1, column=j, value=txt)
                cell.fill = section_fill; cell.font = black_bold; cell.alignment = center; cell.border = border

            k = 0

            for i, row_data in enumerate(content.get("tabla-planAud", []), start=sep7+2):
                ws5.row_dimensions[i].height = 40

                # contador en la columna 2
                k += 1
                cell = ws5.cell(row=i, column=2, value=k)
                cell.alignment = wrap; cell.border = border; cell.alignment = center

                # resto de datos desde la columna 3
                for j, val in enumerate(row_data, start=3):
                    cell = ws5.cell(row=i, column=j, value=val)
                    cell.alignment = wrap; cell.border = border; cell.alignment = center

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=audit_plan_{plan.id}.xlsx'
        return response

    except AuditPlan.DoesNotExist:
        return JsonResponse({"error": "Plan no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def download_excel_report(request):
    
    if request.user.role not in ['auditLeaderUser', 'superUser', 'organizationUser']:
        return JsonResponse({"error": "No autorizado"}, status=403)

    report_id = request.GET.get("report_id")
    if not report_id:
        return JsonResponse({"error": "ID de report no proporcionado"}, status=400)

    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, RadarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import SeriesLabel
    
    try:
        report = Report.objects.get(id=report_id)

        resumen_raw = json.loads(report.resumen_data or "[]")
        resumen_headers = ["Ciclo de vida", "Proceso", "Fortalezas", "Recomendaciones", "Riesgos", "No Conformidades", "Total"]
        resumen = [dict(zip(resumen_headers, row)) for row in resumen_raw]

        detail_cols = ["No", "Cláusula", "Norma", "Hallazgo", "Evidencia", "Dependencia", "Lugar", "Proceso", "Tipo Proceso"]
        
        sheets_data = {}
        for name, data_str in [
            ("Fortalezas", report.fortalezas_data),
            ("Conformidades", report.conformidades_data),
            ("Recomendaciones", report.recomendaciones_data),
            ("Riesgos", report.riesgos_data),
            ("NoConformidades", report.no_conformidades_data),
        ]:
            data = json.loads(data_str or "[]")
            formatted = [dict(zip(detail_cols, row)) for row in data]
            sheets_data[name] = formatted

        # ✅ ESTILOS MEJORADOS
        thin = Side(border_style="thin", color="000000")
        medium = Side(border_style="medium", color="000000")
        border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)
        
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
        wrap_top = Alignment(wrap_text=True, vertical="top")
        
        # Colores corporativos
        header_fill = PatternFill("solid", fgColor="1F4E78")  # Azul oscuro
        section_fill = PatternFill("solid", fgColor="4472C4")  # Azul medio
        alt_row_fill = PatternFill("solid", fgColor="D9E1F2")  # Azul claro
        
        white_bold = Font(color="FFFFFF", bold=True, size=12)
        black_bold = Font(color="000000", bold=True, size=11)
        normal_font = Font(color="000000", size=10)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            wb = writer.book

            # ========================================
            # 1) HOJA RESUMEN
            # ========================================
            ws = wb.create_sheet("Resumen", 0)
            
            # Anchos de columna
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            for col in ['C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[col].width = 15

            row = 1
            
            # Sección de criterios (Pertinencia, Adecuación, Eficacia)
            criterios = [
                ("Pertinencia", report.pertinencia_data),
                ("Adecuación", report.adecuacion_data),
                ("Eficacia", report.eficacia_data),
            ]
            
            for titulo, contenido in criterios:
                # Título
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                c = ws.cell(row=row, column=1, value=titulo)
                c.fill = header_fill
                c.font = white_bold
                c.alignment = center
                c.border = border_medium
                ws.row_dimensions[row].height = 30
                row += 1

                # Contenido
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
                c = ws.cell(row=row, column=1, value=contenido)
                c.alignment = left_wrap
                c.border = border_thin
                c.font = normal_font
                ws.row_dimensions[row].height = 80
                row += 2

            # ✅ TABLA RESUMEN GENERAL
            row += 1
            table_start_row = row
            
            headers = ["Ciclo de vida", "Proceso", "Fortalezas", "Recomendaciones", "Riesgos", "No Conformidades", "Total"]
            for j, txt in enumerate(headers, start=1):
                c = ws.cell(row=row, column=j, value=txt)
                c.fill = section_fill
                c.font = white_bold
                c.alignment = center
                c.border = border_medium
            ws.row_dimensions[row].height = 25

            # Datos de la tabla
            resumen_data = json.loads(report.resumen_data or "[]")
            data_start_row = row + 1
            
            for idx, resumen_row in enumerate(resumen_data):
                row += 1
                
                # Alternar colores de fila
                row_fill = alt_row_fill if idx % 2 == 0 else None
                
                for j, val in enumerate(resumen_row[:6], start=1):
                    if j == 2:  # Segundo elemento — se omite
                        continue
                    if j >= 3:  # Columnas numéricas
                        try:
                            val = float(val)
                        except:
                            pass
                    
                    c = ws.cell(row=row, column=j, value=val)
                    c.alignment = center if j >= 3 else left_wrap
                    c.border = border_thin
                    c.font = normal_font
                    if row_fill:
                        c.fill = row_fill

                # Columna Total con fórmula
                total_formula = f"=SUM(C{row}:F{row})"
                c = ws.cell(row=row, column=7, value=total_formula)
                c.alignment = center
                c.border = border_thin
                c.font = black_bold
                if row_fill:
                    c.fill = row_fill
                    
                ws.row_dimensions[row].height = 25

            data_end_row = row

            # ✅ FILA DE TOTALES
            row += 1
            ws.cell(row=row, column=1, value="TOTAL").font = black_bold
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="FFC000")
            ws.cell(row=row, column=1).border = border_medium
            
            for col in range(3, 8):  # Columnas C-G
                letra_col = get_column_letter(col)
                formula = f"=SUM({letra_col}{data_start_row}:{letra_col}{data_end_row})"
                c = ws.cell(row=row, column=col, value=formula)
                c.font = black_bold
                c.alignment = center
                c.fill = PatternFill("solid", fgColor="FFC000")
                c.border = border_medium
            
            totals_row = row

            # ========================================
            # 2) GRÁFICAS
            # ========================================
            
            # ✅ GRÁFICA 1: Diagrama de Líneas (estilo de tu imagen)
            line_chart = LineChart()
            line_chart.title = "Cantidad de Hallazgos"
            line_chart.style = 12
            line_chart.y_axis.title = "Cantidad"
            line_chart.x_axis.title = "Ciclo de Vida"
            line_chart.height = 10
            line_chart.width = 20
            
            # Datos: Fortalezas, Recomendaciones, Riesgos, No Conformidades
            cats = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
            
            for col, title in [(3, "FOR"), (4, "REC"), (5, "RIE"), (6, "DNC")]:
                data = Reference(ws, min_col=col, min_row=data_start_row-1, max_row=data_end_row)
                line_chart.add_data(data, titles_from_data=True)
            
            line_chart.set_categories(cats)
            ws.add_chart(line_chart, f"I{table_start_row}")

            # ✅ GRÁFICA 2: Diagrama de Barras - Ciclo de Vida vs Habilitadores
            # Separar datos por tipo
            ciclo_vida_count = 0
            habilitadores_count = 0
            
            ciclo_de_vida = []
            habilitadores = []

            for resumen_row in resumen_data:
                tipo_proceso = str(resumen_row[0]).lower() if len(resumen_row) > 0 else ""
                if "ciclo de vida" in tipo_proceso or "ciclo" in tipo_proceso:
                    ciclo_vida_count += 1
                    ciclo_de_vida.append(tipo_proceso)
                elif "habilitador" in tipo_proceso:
                    habilitadores_count += 1
                    habilitadores.append(tipo_proceso)  
                
            rows = [
                ['Fortalezas', 'Conformidades', 'Recomendaciones', 'Riesgos', 'No Conformidades'],
                    
            ]

            # Crear datos para la gráfica
            pie_data_row = data_end_row + 3
            ws.cell(row=pie_data_row, column=9, value="Tipo")
            ws.cell(row=pie_data_row, column=10, value="Cantidad")
            ws.cell(row=pie_data_row + 1, column=9, value="Ciclo de Vida")
            ws.cell(row=pie_data_row + 1, column=10, value=ciclo_vida_count)
            ws.cell(row=pie_data_row + 2, column=9, value="Habilitadores")
            ws.cell(row=pie_data_row + 2, column=10, value=habilitadores_count)
            
            # Gráfica de pastel
            pie_chart = PieChart()
            pie_chart.title = "Distribución Ciclo de Vida vs Habilitadores"
            pie_chart.height = 10
            pie_chart.width = 12
            
            labels = Reference(ws, min_col=9, min_row=pie_data_row+1, max_row=pie_data_row+2)
            data = Reference(ws, min_col=10, min_row=pie_data_row, max_row=pie_data_row+2)
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(labels)
            
            # Mostrar porcentajes
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True
            
            ws.add_chart(pie_chart, f"I{table_start_row + 18}")

            # ✅ GRÁFICA 3: Riesgos y No Conformidades
            risk_data_row = pie_data_row + 4
            
            # Sumar Riesgos y No Conformidades por tipo de proceso
            ciclo_riesgos = 0
            ciclo_nc = 0
            hab_riesgos = 0
            hab_nc = 0
            
            for resumen_row in resumen_data:
                if len(resumen_row) < 6:
                    continue
                    
                tipo_proceso = str(resumen_row[0]).lower()
                try:
                    riesgos = float(resumen_row[4]) if len(resumen_row) > 4 else 0
                    no_conf = float(resumen_row[5]) if len(resumen_row) > 5 else 0
                except:
                    riesgos = 0
                    no_conf = 0
                
                if "ciclo de vida" in tipo_proceso or "ciclo" in tipo_proceso:
                    ciclo_riesgos += riesgos
                    ciclo_nc += no_conf
                elif "habilitador" in tipo_proceso:
                    hab_riesgos += riesgos
                    hab_nc += no_conf
            
            # Crear tabla de datos
            ws.cell(row=risk_data_row, column=9, value="Tipo")
            ws.cell(row=risk_data_row, column=10, value="Riesgos")
            ws.cell(row=risk_data_row, column=11, value="No Conformidades")
            
            ws.cell(row=risk_data_row + 1, column=9, value="Ciclo de Vida")
            ws.cell(row=risk_data_row + 1, column=10, value=ciclo_riesgos)
            ws.cell(row=risk_data_row + 1, column=11, value=ciclo_nc)
            
            ws.cell(row=risk_data_row + 2, column=9, value="Habilitadores")
            ws.cell(row=risk_data_row + 2, column=10, value=hab_riesgos)
            ws.cell(row=risk_data_row + 2, column=11, value=hab_nc)
            
            # Gráfica de barras
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.title = "Riesgos y No Conformidades"
            bar_chart.style = 10
            bar_chart.y_axis.title = "Cantidad"
            bar_chart.x_axis.title = "Tipo de Proceso"
            bar_chart.height = 10
            bar_chart.width = 12
            
            cats = Reference(ws, min_col=9, min_row=risk_data_row+1, max_row=risk_data_row+2)
            data = Reference(ws, min_col=10, max_col=11, min_row=risk_data_row, max_row=risk_data_row+2)
            
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)
            
            ws.add_chart(bar_chart, f"I{table_start_row + 36}")

            # ✅ GRÁFICA 4: Total General (Barras)
            total_chart = BarChart()
            total_chart.type = "col"
            total_chart.title = "Hallazgos por Ciclo de Vida"
            total_chart.style = 11
            total_chart.y_axis.title = "Cantidad"
            total_chart.x_axis.title = "Ciclo de Vida"
            total_chart.height = 10
            total_chart.width = 12

            # Categorías: col A, solo filas de datos (sin header)
            cats_total = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)

            # Datos: cols C-F, mismo rango de filas que las categorías (sin header)
            data_total = Reference(ws, min_col=3, max_col=6, min_row=data_start_row, max_row=data_end_row)

            total_chart.add_data(data_total, titles_from_data=False)
            total_chart.set_categories(cats_total)

            # Títulos de series asignados manualmente para evitar referencias cruzadas
            for i, title in enumerate(["Fortalezas", "Recomendaciones", "Riesgos", "No Conformidades"]):
                total_chart.series[i].title = SeriesLabel(v=title)

            ws.add_chart(total_chart, f"I{table_start_row + 54}")

            # ========================================
            # 3) HOJAS DE DETALLE
            # ========================================
            for name, data in sheets_data.items():
                ws2 = wb.create_sheet(name)
                
                # Anchos de columna
                ws2.column_dimensions['A'].width = 8
                ws2.column_dimensions['B'].width = 12
                ws2.column_dimensions['C'].width = 15
                ws2.column_dimensions['D'].width = 40
                ws2.column_dimensions['E'].width = 25
                ws2.column_dimensions['F'].width = 15
                ws2.column_dimensions['G'].width = 15
                ws2.column_dimensions['H'].width = 20
                ws2.column_dimensions['I'].width = 20

                # Título de la hoja
                ws2.merge_cells("A1:I1")
                c = ws2["A1"]
                c.value = name.upper()
                c.fill = header_fill
                c.font = Font(color="FFFFFF", bold=True, size=14)
                c.alignment = center
                c.border = border_medium
                ws2.row_dimensions[1].height = 35

                # Fila vacía
                ws2.row_dimensions[2].height = 5
                
                # Encabezados
                for j, h in enumerate(detail_cols, start=1):
                    c = ws2.cell(row=3, column=j, value=h)
                    c.fill = section_fill
                    c.font = white_bold
                    c.alignment = center
                    c.border = border_medium
                ws2.row_dimensions[3].height = 25

                # Datos
                for i, record in enumerate(data, start=4):
                    # Alternar colores
                    row_fill = alt_row_fill if (i - 4) % 2 == 0 else None
                    
                    ws2.row_dimensions[i].height = 30
                    for j, key in enumerate(detail_cols, start=1):
                        val = record.get(key, "")
                        c = ws2.cell(row=i, column=j, value=val)
                        c.alignment = center if j == 1 else left_wrap
                        c.border = border_thin
                        c.font = normal_font
                        if row_fill:
                            c.fill = row_fill

            # ========================================
            # 4) HOJA DATOS GENERALES
            # ========================================
            meta = {
                "ID Informe":      report.id,
                "Fecha":           report.creation_date.strftime('%Y-%m-%d'),
                "Organización":    report.organization.name,
                "Líder Auditor":   report.leader_auditor.get_full_name(),
                "Total Cláusulas": report.total_clauses(),
            }
            
            ws3 = wb.create_sheet("Datos Generales")
            
            # Encabezados
            for i, key in enumerate(meta.keys(), start=1):
                ws3.column_dimensions[get_column_letter(i)].width = 25
                cell = ws3.cell(row=1, column=i, value=key)
                cell.fill = section_fill
                cell.font = white_bold
                cell.alignment = center
                cell.border = border_medium
                
                # Valores
                cell = ws3.cell(row=2, column=i, value=meta[key])
                cell.alignment = center
                cell.border = border_thin
                cell.font = normal_font
                
            ws3.row_dimensions[1].height = 30
            ws3.row_dimensions[2].height = 25

            # Eliminar hoja por defecto si existe
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        resp = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="informe_auditoria_{report.id}.xlsx"'
        return resp

    except Report.DoesNotExist:
        return JsonResponse({"error": "Informe no encontrado"}, status=404)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=500)